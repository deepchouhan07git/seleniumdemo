import openpyxl

file="C:\\Users\\user\\Downloads\\Book1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active # or workbook["Data"] get the sheet

# Loop to add content in the excel sheet
# for r in range(1,6):
#     for c in range(1,5):
#         sheet.cell(r,c).value=""

# For different types of data

sheet.cell(1,1).value="Emp ID"
sheet.cell(1,2).value="Emp Name"
sheet.cell(1,3).value="Employer"

sheet.cell(2,1).value="545"
sheet.cell(2,2).value="Smith"
sheet.cell(2,3).value="HCL"

sheet.cell(3,1).value="874"
sheet.cell(3,2).value="John"
sheet.cell(3,3).value="ACC"

workbook.save(file)


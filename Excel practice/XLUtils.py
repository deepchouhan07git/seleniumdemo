import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file, sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return (sheet.max_row)

def getcolCount(file, sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return (sheet.max_column)

def readData(file, sheetname, rowno, columno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(rowno, columno).value

def writeData(file, sheetname, rowno, columno, data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(rowno, columno).value=data
    workbook.save(file)

def fillGreencolor(file, sheetname, rowno, columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    greenFill=PatternFill(start_color='60b212',
                end_color='60b212',
                fill_type='solid')
    sheet.cell(rowno,columno).fill=greenFill
    workbook.save(file)

def fillRedcolor(file, sheetname, rowno, columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    redFill=PatternFill(start_color='ff0000',
                end_color='ff0000',
                fill_type='solid')
    sheet.cell(rowno,columno).fill=redFill
    workbook.save(file)
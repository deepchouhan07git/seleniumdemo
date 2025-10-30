import openpyxl

# file --> workbook --> sheets --> rows -- > cells

file="C:\\users\\user\\Downloads\\Semi_Automatic_Car_Wash_Budget.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Setup Costs"]

rows=sheet.max_row # Read no. of rows in a sheet
cols=sheet.max_column # Read no. of columns in a sheet

# Loop to get all the data
for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(r,c).value,end="               ")
    print()